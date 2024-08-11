import sys
import os
import qdarkstyle
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from PyQt5.QtWidgets import (
    QVBoxLayout, QWidget, QLabel, QPushButton, 
    QLineEdit, QDateEdit, QApplication, QMainWindow, QMessageBox, QTabWidget, QSlider
)
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtCore import QDate, QSize, Qt
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.dates as mdates
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backend_bases import PickEvent
from PyQt5.QtWidgets import  QComboBox
from PyQt5.QtGui import QFont
import pandas as pd

def plot_expense_summary(file_name, ax):
    # Load the data from the Excel file
    df = pd.read_excel(file_name)

    # Ensure 'Amount' is numeric
    df['Amount'] = df['Amount'].replace(',', '', regex=True).astype(float)

    # Group by 'Expense Type' and sum the 'Amount'
    expense_summary = df.groupby('Expense Type')['Amount'].sum().reset_index()

    # Clear the previous plot
    ax.clear()
    
    # Set dark mode style
    ax.set_facecolor('#2E2E2E')
    ax.figure.set_facecolor('#2E2E2E')

    # Create a bar plot with white labels and dark background
    bars = ax.bar(expense_summary['Expense Type'], expense_summary['Amount'], color='#4CAF50')

    # Set label colors to white
    ax.set_xlabel('Expense Type', color='white')
    ax.set_ylabel('Total Amount', color='white')
    ax.set_title('Total Expenditure by Expense Type', color='white')

    # Set tick parameters to white
    ax.tick_params(axis='x', colors='white')
    ax.tick_params(axis='y', colors='white')

    # Add value labels on top of the bars
    for bar in bars:
        yval = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, yval, f'{yval:.2f}', ha='center', va='bottom', color='white')

    # Rotate x-axis labels for better readability
    ax.tick_params(axis='x', rotation=45)

    # Adjust layout
    ax.figure.tight_layout()

class SettingsPage(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        self.savings_label = QLabel("Savings (50%): 50%")
        self.savings_slider = self.create_slider(0.0, 1.0, 0.5, self.savings_label)
        self.savings_slider.valueChanged.connect(lambda: self.update_label(self.savings_slider, self.savings_label, "Savings"))

        self.expenditure_label = QLabel("Expenditure (20%): 20%")
        self.expenditure_slider = self.create_slider(0.0, 1.0, 0.2, self.expenditure_label)
        self.expenditure_slider.valueChanged.connect(lambda: self.update_label(self.expenditure_slider, self.expenditure_label, "Expenditure"))

        self.pleasure_label = QLabel("Pleasure (10%): 10%")
        self.pleasure_slider = self.create_slider(0.0, 1.0, 0.1, self.pleasure_label)
        self.pleasure_slider.valueChanged.connect(lambda: self.update_label(self.pleasure_slider, self.pleasure_label, "Pleasure"))

        self.parents_label = QLabel("Parents (10%): 10%")
        self.parents_slider = self.create_slider(0.0, 1.0, 0.1, self.parents_label)
        self.parents_slider.valueChanged.connect(lambda: self.update_label(self.parents_slider, self.parents_label, "Parents"))

        self.gifts_label = QLabel("Gifts (10%): 10%")
        self.gifts_slider = self.create_slider(0.0, 1.0, 0.1, self.gifts_label)
        self.gifts_slider.valueChanged.connect(lambda: self.update_label(self.gifts_slider, self.gifts_label, "Gifts"))

        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.apply_percentages)
        self.ok_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border-radius: 20px;
                padding: 10px 20px;
                font-size: 16px;
                border: none;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.ok_button.setFixedSize(200, 50)

        layout.addWidget(self.savings_label)
        layout.addWidget(self.savings_slider)
        layout.addWidget(self.expenditure_label)
        layout.addWidget(self.expenditure_slider)
        layout.addWidget(self.pleasure_label)
        layout.addWidget(self.pleasure_slider)
        layout.addWidget(self.parents_label)
        layout.addWidget(self.parents_slider)
        layout.addWidget(self.gifts_label)
        layout.addWidget(self.gifts_slider)
        layout.addWidget(self.ok_button)

        self.setLayout(layout)

    def create_slider(self, min_val, max_val, default_val, label):
        min_int = int(min_val * 100)
        max_int = int(max_val * 100)
        default_int = int(default_val * 100)

        slider = QSlider()
        slider.setMinimum(min_int)
        slider.setMaximum(max_int)
        slider.setValue(default_int)
        slider.setOrientation(Qt.Horizontal)
        slider.setTickPosition(QSlider.TicksBelow)
        slider.setTickInterval(10)
        slider.setSingleStep(1)
        slider.setStyleSheet("""
            QSlider::handle:horizontal {
                background: #4CAF50;
                border-radius: 10px;
                height: 20px;
                width: 20px;
                margin: -10px 0;
                border: 1px solid #4CAF50;
            }
            QSlider::groove:horizontal {
                background: #2E2E2E;
                height: 8px;
                border-radius: 4px;
            }
        """)
        return slider

    def update_label(self, slider, label, category):
        value = slider.value() / 100.0
        label.setText(f"{category} ({value * 100:.1f}%): {value * 100:.1f}%")

    def apply_percentages(self):
        savings_percent = self.get_slider_value(self.savings_slider)
        expenditure_percent = self.get_slider_value(self.expenditure_slider)
        pleasure_percent = self.get_slider_value(self.pleasure_slider)
        parents_percent = self.get_slider_value(self.parents_slider)
        gifts_percent = self.get_slider_value(self.gifts_slider)

        total_percentage = savings_percent + expenditure_percent + pleasure_percent + parents_percent + gifts_percent

        if abs(total_percentage - 1.0) > 0.01:
            QMessageBox.warning(self, "Invalid Percentages", "The total percentage must be 100%.")
            return

        QMessageBox.information(self, "Success", "Percentages set successfully!")

    def get_slider_value(self, slider):
        return slider.value() / 100.0





class PieChartPage(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # Use QVBoxLayout to stack widgets vertically
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

    def plot_category_distribution(self, file_name):
        try:
            # Check if file exists
            if not os.path.exists(file_name):
                raise FileNotFoundError("The specified file does not exist.")
            
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active

            categories = ['Savings', 'Expenditure', 'Pleasure', 'Parents', 'Gifts']
            amounts = {category: 0 for category in categories}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 8:
                    raise ValueError("Data row is missing values.")
                try:
                    amounts['Savings'] += row[3] if row[3] is not None else 0
                    amounts['Expenditure'] += row[4] if row[4] is not None else 0
                    amounts['Pleasure'] += row[5] if row[5] is not None else 0
                    amounts['Parents'] += row[6] if row[6] is not None else 0
                    amounts['Gifts'] += row[7] if row[7] is not None else 0
                except (IndexError, TypeError) as e:
                    pass
                    raise ValueError(f"Error processing row data: {e}")

            self.figure = plt.figure(figsize=(12, 6), tight_layout=True)  # Increased figure size
            self.ax = self.figure.add_subplot(111)

            wedges, texts, autotexts = self.ax.pie(
                amounts.values(), 
                labels=categories, 
                autopct=lambda p: '{:.0f}'.format(p * sum(amounts.values()) / 100), 
                startangle=140,
                colors=plt.get_cmap('tab10').colors
            )
            self.ax.set_title('Category Distribution', fontsize=18, fontweight='bold', color='white')  # Title color white

            # Apply dark background and text colors
            self.figure.patch.set_facecolor('#1E1E1E')
            self.ax.set_facecolor('#2E2E2E')
            for text in texts:
                text.set_color('white')
            for autotext in autotexts:
                autotext.set_color('white')

            self.canvas = FigureCanvas(self.figure)
            self.canvas.setFixedHeight(400)  # Increased canvas height

            # Remove all widgets from the layout
            while self.layout.count():
                item = self.layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()

            # Add the canvas to the layout
            self.layout.addWidget(self.canvas)
            self.canvas.draw()

       
        except Exception as e:
             QMessageBox.information(self,":-)","Press ok to create a new record")




class ExpenditurePage(QWidget):
    def __init__(self):
        super().__init__()
        self.file_name = self.generate_file_name()  # Generate the file name with the month's name
        self.ensure_file_exists()  # Ensure the file exists
        self.initUI()
        
    def generate_file_name(self):
        # Get the current month's name and year to include in the file name
        month_name = datetime.now().strftime("%B")  # Full month name (e.g., "August")
        year = datetime.now().strftime("%Y")  # Year (e.g., "2024")
        return f"expenditure_{month_name}_{year}.xlsx"

    def initUI(self):
        layout = QVBoxLayout()

        # Expenditure Amount Input
        self.expenditure_label = QLabel("Expenditure Amount:")
        self.expenditure_input = QLineEdit()
        self.expenditure_input.setPlaceholderText("Enter expenditure amount")
        self.expenditure_input.setFont(QFont("Arial", 12))
        self.expenditure_input.setFixedHeight(40)

        # Expense Type Dropdown
        self.expense_type_label = QLabel("Expense Type:")
        self.expense_type_dropdown = QComboBox()
        self.expense_type_dropdown.addItems([
            "Expenditure",
            "Gifts",
            "Parents",
            "Pleasure"
        ])
        self.expense_type_dropdown.setFont(QFont("Arial", 12))
        self.expense_type_dropdown.setFixedHeight(40)

        # Details Input
        self.details_label = QLabel("Details:")
        self.details_input = QLineEdit()
        self.details_input.setPlaceholderText("Enter details")
        self.details_input.setFont(QFont("Arial", 12))
        self.details_input.setFixedHeight(40)

        # Add Expenditure Button
        self.add_expenditure_button = QPushButton("Add Expenditure")
        self.add_expenditure_button.clicked.connect(self.add_expenditure)
        self.add_expenditure_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border-radius: 20px;
                padding: 10px 20px;
                font-size: 16px;
                border: none;
            }
            QPushButton:hover {
                background-color: #e53935;
            }
        """)
        self.add_expenditure_button.setFixedSize(200, 50)

        # Matplotlib Figure and Canvas
        self.figure, self.ax = plt.subplots(figsize=(10, 6))
        self.canvas = FigureCanvas(self.figure)
        
        layout.addWidget(self.expenditure_label)
        layout.addWidget(self.expenditure_input)
        layout.addWidget(self.expense_type_label)
        layout.addWidget(self.expense_type_dropdown)
        layout.addWidget(self.details_label)
        layout.addWidget(self.details_input)
        layout.addWidget(self.add_expenditure_button)
        layout.addWidget(self.canvas)  # Add the plot widget to the layout

        self.setLayout(layout)
        self.update_plot()  # Initialize the plot

    def ensure_file_exists(self):
        if not os.path.isfile(self.file_name):
            # Create a new workbook and add headers
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Date", "Expense Type", "Amount", "Details"])
            wb.save(self.file_name)

    def add_expenditure(self):
        amount_text = self.expenditure_input.text()
        expense_type = self.expense_type_dropdown.currentText()
        details = self.details_input.text()

        try:
            # Remove commas from amount_text and convert to float
            amount = float(amount_text.replace(',', ''))
            if amount <= 0:
                raise ValueError("Amount must be positive.")
        except ValueError as e:
            QMessageBox.warning(self, "Invalid Input", f"Please enter a valid expenditure amount. Error: {e}")
            return

        try:
            wb = openpyxl.load_workbook(self.file_name)
            ws = wb.active

            # Append expenditure details to the file
            ws.append([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                expense_type,  # Expense type
                f"{amount:,.2f}",  # Format amount with commas
                details  # Details
            ])

            wb.save(self.file_name)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred while saving the expenditure. Error: {e}")
            return

        self.expenditure_input.clear()
        self.details_input.clear()
        QMessageBox.information(self, "Success", "Expenditure added successfully!")

        # Update the plot after adding a new expenditure
        self.update_plot()

    def update_plot(self):
        plot_expense_summary(self.file_name, self.ax)
        self.canvas.draw()  # Redraw the canvas with the updated plot



class EntryGraphPage(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)
    
    def clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

    def plot_entries(self, file_name):
        
        try:
            # Check if file exists
            if not os.path.exists(file_name):
                raise FileNotFoundError("The specified file does not exist.")
            
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active

            # Initialize lists to store data
            sources = []
            savings = []
            expenditure = []
            pleasure = []
            parents = []
            gifts = []

            # Read data from Excel file
            for row in ws.iter_rows(min_row=2, values_only=True):
                source = row[1]  # Source of money
                sources.append(source)
                savings.append(row[3] if row[3] is not None else 0)
                expenditure.append(row[4] if row[4] is not None else 0)
                pleasure.append(row[5] if row[5] is not None else 0)
                parents.append(row[6] if row[6] is not None else 0)
                gifts.append(row[7] if row[7] is not None else 0)

            # Create figure and axes
            self.figure, self.ax = plt.subplots(figsize=(12, 8), tight_layout=True)

            # Set width for each bar group
            bar_width = 0.15
            index = range(len(sources))

            # Plot bars for each category with updated colors and transparency
            bars_savings = self.ax.bar([i - 2*bar_width for i in index], savings, bar_width, label='Savings (50%)', color='#6baed6', alpha=0.7, picker=True)  # Light Blue
            bars_expenditure = self.ax.bar([i - bar_width for i in index], expenditure, bar_width, label='Expenditure (20%)', color='#fdae6b', alpha=0.7, picker=True)  # Light Orange
            bars_pleasure = self.ax.bar(index, pleasure, bar_width, label='Pleasure (10%)', color='#a1d76a', alpha=0.7, picker=True)  # Light Green
            bars_parents = self.ax.bar([i + bar_width for i in index], parents, bar_width, label='Parents (10%)', color='#f46d43', alpha=0.7, picker=True)  # Light Red
            bars_gifts = self.ax.bar([i + 2*bar_width for i in index], gifts, bar_width, label='Gifts (10%)', color='#9e9ac8', alpha=0.7, picker=True)  # Light Purple

            # Set labels, title, and ticks
            self.ax.set_xlabel('Source')
            self.ax.set_ylabel('Amount')
            self.ax.set_title('Amounts by Source and Category')
            self.ax.set_xticks(index)
            self.ax.set_xticklabels(sources, rotation=45, ha='right')
            self.ax.legend()

            # Apply dark theme
            self.figure.patch.set_facecolor('#1E1E1E')
            self.ax.set_facecolor('#2E2E2E')
            self.ax.tick_params(axis='x', colors='white')
            self.ax.tick_params(axis='y', colors='white')
            self.ax.grid(True, linestyle='--', linewidth=0.7, color='white')

            # Add event handler for clicking on bars
            self.figure.canvas.mpl_connect('pick_event', self.on_pick)

            # Create canvas and add it to the layout
            self.canvas = FigureCanvas(self.figure)
            self.layout.addWidget(self.canvas)
            self.canvas.draw()
        
        

        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred while plotting the data. Error: {e}")
    
    def refresh_plot(self, file_name):
        self.clear_layout(self.layout)
        
        self.figure.clear()
        self.plot_entries(file_name)

    def on_pick(self, event: PickEvent):
        # Check if the event is from a bar
        if isinstance(event.artist, plt.Rectangle):
            bar = event.artist
            height = bar.get_height()
            label = bar.get_label()
            # Show the amount in a message box
            QMessageBox.information(self, "Amount", f"Money: {height:.2f}")





class SavingsTrackerApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Savings Tracker")
        self.setGeometry(100, 100, 1000, 700)
        self.setWindowIcon(QIcon("savings_icon.svg"))

        self.temp_dir = os.getcwd()
        self.create_excel_file()

        self.initUI()
        self.update_total_savings()
        self.plot_recent_expenses()


    def initUI(self):
        layout = QVBoxLayout()

        self.amount_label = QLabel("Amount Received:")
        self.amount_input = QLineEdit()
        self.amount_input.setPlaceholderText("Enter amount received")
        self.amount_input.setFont(QFont("Arial", 12))
        self.amount_input.setFixedHeight(40)

        self.source_label = QLabel("Source of Money:")
        self.source_input = QLineEdit()
        self.source_input.setPlaceholderText("Enter source of money")
        self.source_input.setFont(QFont("Arial", 12))
        self.source_input.setFixedHeight(40)

        self.delivery_date_label = QLabel("Delivery Date:")
        self.delivery_date_input = QDateEdit()
        self.delivery_date_input.setCalendarPopup(True)
        self.delivery_date_input.setFont(QFont("Arial", 12))
        self.delivery_date_input.setFixedHeight(40)
        self.delivery_date_input.setDate(QDate.currentDate())

        self.add_button = QPushButton("Add Entry")
        self.add_button.setIcon(QIcon("add.svg"))
        self.add_button.setIconSize(QSize(24, 24))
        self.add_button.setFixedHeight(50)
        self.add_button.setFont(QFont("Arial", 12, QFont.Bold))
        self.add_button.clicked.connect(self.add_entry)

        self.total_savings_label = QLabel("Total Savings:")
        self.total_savings_display = QLabel("₦0.00")
        self.total_savings_display.setFont(QFont("Arial", 16, QFont.Bold))
        self.total_savings_display.setStyleSheet("color: #4CAF50; padding: 10px; border: 2px solid #4CAF50; border-radius: 5px; background-color: #1E1E1E;")

        self.toggle_button = QPushButton("Show Total Savings")
        self.toggle_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border-radius: 20px;
                padding: 10px 20px;
                font-size: 16px;
                border: none;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.toggle_button.setFixedSize(200, 50)
        self.toggle_button.clicked.connect(self.toggle_total_savings)

        # Tabs
        self.tab_widget = QTabWidget()
        self.main_page = QWidget()

        self.settings_page = SettingsPage()
        self.pie_chart_page = PieChartPage()
        self.entry_graph_page = EntryGraphPage()
        self.expenditure_page = ExpenditurePage()
        self.settings_page = SettingsPage()

        self.tab_widget.addTab(self.main_page, "Main")
        self.tab_widget.addTab(self.pie_chart_page, "Category Distribution")
        self.tab_widget.addTab(self.entry_graph_page, "Category Graph")
        self.tab_widget.addTab(self.expenditure_page, "Add Expenditure")
        self.tab_widget.addTab(self.settings_page, "Settings")
        
        

        

       
        
            
        self.entry_graph_page.plot_entries(self.file_name)

        

        # Main Page Layout
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.amount_label)
        main_layout.addWidget(self.amount_input)
        main_layout.addWidget(self.source_label)
        main_layout.addWidget(self.source_input)
        main_layout.addWidget(self.delivery_date_label)
        main_layout.addWidget(self.delivery_date_input)
        main_layout.addWidget(self.add_button)
        main_layout.addWidget(self.total_savings_label)
        main_layout.addWidget(self.total_savings_display)
        main_layout.addWidget(self.toggle_button)

        # Initialize the canvas for plotting
        self.figure = plt.figure(figsize=(8, 5), tight_layout=True)
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setFixedHeight(300)
        main_layout.addWidget(self.canvas)

        self.main_page.setLayout(main_layout)

        container = QWidget()
        container.setLayout(QVBoxLayout())
        container.layout().addWidget(self.tab_widget)
        self.setCentralWidget(container)

        

    def create_excel_file(self):
        now = datetime.now()
        year = now.strftime("%Y")
        month_name = now.strftime("%B")
        self.file_name = os.path.join(self.temp_dir, f"savings_tracker_{month_name}_{year}.xlsx")

        if not os.path.exists(self.file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Savings Tracker"

            headers = [
                "Date", "Source", "Amount Received", "Savings (50%)", "Expenditure (20%)",
                "Pleasure (10%)", "Parents (10%)", "Gifts (10%)", "Total Savings", "Delivery Date",
                "Savings Total", "Expenditure Total", "Pleasure Total", "Parents Total", "Gifts Total"
            ]
            ws.append(headers)

            wb.save(self.file_name)
            print(f"{self.file_name} created successfully.")

    def add_entry(self):
        amount = self.amount_input.text()
        source = self.source_input.text()
        delivery_date = self.delivery_date_input.date().toString("yyyy-MM-dd")

        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError("Amount must be positive.")
        except ValueError as e:
            QMessageBox.warning(self, "Invalid Input", "Please enter a valid amount")
            return

        if not source:
            QMessageBox.warning(self, "Missing Source", "Please enter the source of the money.")
            return

        savings = self.get_percentage('savings') * amount
        expenditure = self.get_percentage('expenditure') * amount
        pleasure = self.get_percentage('pleasure') * amount
        parents = self.get_percentage('parents') * amount
        gifts = self.get_percentage('gifts') * amount
        total_savings = savings + expenditure + pleasure + parents + gifts

        try:
            wb = openpyxl.load_workbook(self.file_name)
            ws = wb.active

            last_row = ws.max_row

            if last_row > 1:
                savings_total = ws.cell(row=last_row, column=11).value
                expenditure_total = ws.cell(row=last_row, column=12).value
                pleasure_total = ws.cell(row=last_row, column=13).value
                parents_total = ws.cell(row=last_row, column=14).value
                gifts_total = ws.cell(row=last_row, column=15).value
            else:
                savings_total = 0
                expenditure_total = 0
                pleasure_total = 0
                parents_total = 0
                gifts_total = 0

            # Calculate new totals
            new_savings_total = savings_total + savings
            new_expenditure_total = expenditure_total + expenditure
            new_pleasure_total = pleasure_total + pleasure
            new_parents_total = parents_total + parents
            new_gifts_total = gifts_total + gifts

            ws.append([
                datetime.now().strftime("%Y-%m-%d"),
                source,
                amount,
                savings,
                expenditure,
                pleasure,
                parents,
                gifts,
                total_savings,
                delivery_date,
                new_savings_total,
                new_expenditure_total,
                new_pleasure_total,
                new_parents_total,
                new_gifts_total
            ])

            wb.save(self.file_name)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred while saving the entry. Error: {e}")
            return

        self.amount_input.clear()
        self.source_input.clear()
        self.delivery_date_input.setDate(QDate.currentDate())

        self.update_total_savings()
        self.plot_recent_expenses()
        self.pie_chart_page.plot_category_distribution(self.file_name)
        self.entry_graph_page.refresh_plot(self.file_name)
        self.update_total_savings()

        QMessageBox.information(self, "Success", "Entry added successfully!")

    def update_total_savings(self):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active

        last_row = ws.max_row
        if last_row > 1:
            total_savings = ws.cell(row=last_row, column=11).value
            self.total_savings_display.setText(f"₦{total_savings:,.2f}")

  


    def plot_recent_expenses(self):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active

        dates_sources = []
        total_savings = []
        date_to_color = {}

        # Use a colormap with sufficient unique colors
        colors = plt.cm.tab20.colors

        # Initialize color index
        color_index = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            date_value = row[0]  # Date column
            source = row[1]  # Source column

            if isinstance(date_value, str):
                try:
                    # Convert string to datetime object
                    date_value = datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError:
                    continue  # Skip rows with invalid date formats

            # Format the date string consistently
            date_str = date_value.strftime("%b %d")  # Example: "Aug 10"

            dates_sources.append(f"{date_str},{source}")
            total_savings.append(row[8])  # Total Savings column

            # Assign a color to each date if not already assigned
            if date_str not in date_to_color:
                date_to_color[date_str] = colors[color_index % len(colors)]
                color_index += 1

        self.figure.clear()

        # Extract colors for each date
        bar_colors = [date_to_color[date.split(',')[0]] for date in dates_sources]

        ax1 = self.figure.add_subplot(121)
        ax1.bar(
            dates_sources,
            total_savings,
            color=bar_colors,  # Use the list of colors for each bar
            edgecolor='black'
        )
        ax1.set_xlabel('Source', fontsize=12, fontweight='bold', color='white')
        ax1.set_ylabel('Cash Inflow', fontsize=12, fontweight='bold', color='white')
        ax1.set_title(f'Savings Tracker for {datetime.now().strftime("%B %Y")}', fontsize=14, fontweight='bold', color='white')

        # Rotate x-axis labels for better readability
        ax1.set_xticklabels(dates_sources, rotation=45, ha='right', fontsize=10, fontweight='bold', color='white')

        # Style y-axis
        ax1.yaxis.set_tick_params(labelsize=10, colors='white')
        ax1.xaxis.set_tick_params(labelsize=10, colors='white')

        # Add gridlines for better readability
        ax1.yaxis.grid(True, linestyle='--', linewidth=0.7, color='white')

        # Apply dark background style
        self.figure.patch.set_facecolor('#1E1E1E')
        ax1.set_facecolor('#2E2E2E')

        # Plot category distribution on separate page
        self.pie_chart_page.plot_category_distribution(self.file_name)

        self.canvas.draw()

    
    def toggle_total_savings(self):
        if self.total_savings_display.isVisible():
            self.total_savings_display.setVisible(False)
            self.toggle_button.setText("Show Total Savings")
        else:
            self.total_savings_display.setVisible(True)
            self.toggle_button.setText("Hide Total Savings")

    def get_percentage(self, category):
        settings = self.settings_page
        if category == 'savings':
            return settings.get_slider_value(settings.savings_slider)
        elif category == 'expenditure':
            return settings.get_slider_value(settings.expenditure_slider)
        elif category == 'pleasure':
            return settings.get_slider_value(settings.pleasure_slider)
        elif category == 'parents':
            return settings.get_slider_value(settings.parents_slider)
        elif category == 'gifts':
            return settings.get_slider_value(settings.gifts_slider)
        return 0



if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    window = SavingsTrackerApp()
    window.show()
    sys.exit(app.exec_())
